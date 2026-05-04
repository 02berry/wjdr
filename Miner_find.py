import pandas as pd
import re
import os
import glob
import warnings
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')


# ==================== 工具函数 ====================
def safe_to_numeric(series, default_value=0):
    result = pd.to_numeric(series, errors='coerce')
    return result.fillna(default_value).astype(float)


def extract_furnace_level(level_str):
    if pd.isna(level_str) or str(level_str).strip() == '':
        return 0
    level_str = str(level_str).strip()
    match = re.search(r'lv\.?\s*(\d+)', level_str, re.IGNORECASE)
    if match:
        level = int(match.group(1))
        return level if level <= MAX_LEVEL else 0
    return 0


def has_collecting(troop_str):
    if pd.isna(troop_str) or str(troop_str).strip() == '':
        return False
    return '采集资源' in str(troop_str)


def count_collecting_teams(troop_str):
    if pd.isna(troop_str) or str(troop_str).strip() == '':
        return 0
    return str(troop_str).count('采集资源')


def extract_date_from_filename(filename):
    """
    从文件名中提取日期信息
    支持格式：
    - 3957_0428 (一天一次)
    - 3957_0428a, 3957_0428b (一天多次，用abcd区分)
    返回 (datetime对象, 日期字符串, 序号) 用于排序
    """
    # 匹配模式：3957_后面4位数字+可选字母
    match = re.search(r'3957_(\d{4})([a-z]?)', filename)
    if match:
        date_str = match.group(1)  # 0428
        suffix = match.group(2)  # a/b/c 或空字符串

        # 尝试解析日期（假设是当前年份）
        current_year = datetime.now().year
        try:
            month = int(date_str[:2])
            day = int(date_str[2:])
            date_obj = datetime(current_year, month, day)
            return date_obj, date_str, suffix
        except ValueError:
            pass

        # 如果上面失败，尝试6位格式 MMDDYY
        match2 = re.search(r'3957_(\d{6})([a-z]?)', filename)
        if match2:
            date_str = match2.group(1)
            suffix = match2.group(2)
            try:
                year = 2000 + int(date_str[4:6])
                month = int(date_str[:2])
                day = int(date_str[2:4])
                date_obj = datetime(year, month, day)
                return date_obj, date_str, suffix
            except ValueError:
                pass

    return None, None, ''



# ==================== 配置区 ====================
# 【修改文件读取数量】将下面的数字改成你想要读取的最近文件数量
READ_FILE_COUNT = 999  # 读取最近的文件数量，例如改为5则读取最新5个文件
MIN_RECENT_MINING = 5  # 最近N次采集中最少挖矿次数，低于此次数标蓝（退游矿工）
N_LATEST_DAYS = 5  # 最新N天用于计算"最新挖矿天数"列
LATEST_THRESHOLD = 0.6  # 最新挖矿天数比例阈值，≥此值标绿（近期活跃）
LATEST_MIN = 0.2  # 最新挖矿天数最低比例，低于此值即使历史达标也不输出
EXCLUDE_ALLIANCES = ['SSS', '999','UtM']  # 排除的联盟列表，用最新联盟判断

# 其他配置
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER_PATH = os.path.join(CURRENT_DIR, 'data')
FILE_PATTERN = '3957_*.xlsx'  # 文件命名模式：3957_0428.xlsx 或 3957_0428a.xlsx
THRESHOLD = 0.49  # 历史挖矿率阈值，≥此值进入输出列表
MAX_PRESTIGE = 1500
MAX_LEVEL = 25

# ==================== 查找并筛选文件 ====================
all_files = glob.glob(os.path.join(FOLDER_PATH, FILE_PATTERN))

# 过滤掉备注文件本身
all_files = [f for f in all_files if 'miner_remark' not in os.path.basename(f).lower()]

print(f"\n📁 找到 {len(all_files)} 个数据文件")

# 提取文件日期并排序
file_dates = []
for file_path in all_files:
    filename = os.path.basename(file_path)
    date_obj, date_str, suffix = extract_date_from_filename(filename)
    if date_obj:
        # 排序优先级：1.日期 2.后缀字母（同一天内a最早，b其次）
        file_dates.append((file_path, date_obj, date_str, suffix))
    else:
        print(f"  ⚠️ 无法提取日期: {filename}")

# 按日期排序（最新的在前），同一天按后缀字母排序
file_dates.sort(key=lambda x: (x[1], x[3] if x[3] else ''), reverse=False)
file_dates.sort(key=lambda x: x[1], reverse=True)

# 统计实际天数（不同日期的数量）
unique_dates = set(date_str for _, _, date_str, _ in file_dates)
print(f"  📅 涵盖 {len(unique_dates)} 个不同日期")

# 只保留最近的READ_FILE_COUNT天（不是文件数）
sorted_dates = sorted(unique_dates, reverse=True)[:READ_FILE_COUNT]
selected_files = [(fp, do, ds, sf) for fp, do, ds, sf in file_dates if ds in sorted_dates]

print(f"\n✅ 将处理最近 {len(sorted_dates)} 天的数据（共{len(selected_files)}个文件）:")
current_date = None
for file_path, date_obj, date_str, suffix in selected_files:
    display_date = f"{date_str[:2]}-{date_str[2:]}"
    if suffix:
        display_name = f"{display_date} (第{suffix.upper()}次)"
    else:
        display_name = f"{display_date}"

    # 显示日期分组
    if date_str != current_date:
        print(f"  📅 {display_date}:")
        current_date = date_str

    print(f"    📄 {os.path.basename(file_path)}")

if len(selected_files) == 0:
    print("❌ 没有找到符合条件的文件")
    exit()

# ==================== 读取选定文件 ====================
daily_accounts = []
# ⭐ 记录每个账号的最新完整信息（不受条件限制）
account_latest_full_info = {}  # {账号: (日期排序键, info字典)}

for file_path, date_obj, date_str, suffix in selected_files:
    filename = os.path.basename(file_path)

    try:
        df = pd.read_excel(file_path, sheet_name=0, dtype=str)
    except Exception as e:
        print(f"  ❌ 读取失败 {filename}: {e}")
        continue

    if '兵线' not in df.columns:
        print(f"  ⚠️ {filename}: 无'兵线'列，跳过")
        continue

    df['账号'] = safe_to_numeric(df['账号'])
    df['声望'] = safe_to_numeric(df['声望'])
    df['炉子等级'] = df['炉子'].apply(extract_furnace_level)

    today_dict = {}
    suffix_order = ord(suffix) - ord('a') + 1 if suffix else 0
    date_key = (date_obj, suffix_order)

    for _, row in df.iterrows():
        account = int(row['账号'])
        if account == 0:
            continue

        # ⭐ 不管符不符合条件，都记录最新完整信息
        current_info = {
            '昵称': row['昵称'],
            '联盟': row['联盟'],
            '炉子': row['炉子'],
            '炉子等级': int(row['炉子等级']),
            '声望': int(row['声望']),
            '坐标': row.get('坐标', ''),
            '罩子': row.get('罩子', ''),
            '兵线': row.get('兵线', ''),
        }

        if account not in account_latest_full_info:
            account_latest_full_info[account] = (date_key, current_info)
        else:
            if date_key > account_latest_full_info[account][0]:
                account_latest_full_info[account] = (date_key, current_info)

        # 条件过滤：判断是否算作挖矿账号（声望过滤延后到用最新声望判断）
        if row['炉子等级'] < 1 or row['炉子等级'] > MAX_LEVEL:
            continue

        is_mining = has_collecting(row['兵线'])
        teams = count_collecting_teams(row['兵线'])

        today_dict[account] = {
            'is_mining': is_mining,
            'teams': teams,
            'info': current_info,  # 这里也用同样的info
        }

    daily_accounts.append(today_dict)
    mining_count = sum(1 for v in today_dict.values() if v['is_mining'])

    display_date = f"{date_str[:2]}-{date_str[2:]}"
    print(
        f"  {display_date}{'(' + suffix.upper() + ')' if suffix else ''}: {len(df)}行 → {len(today_dict)}人, 挖矿{mining_count}人")

# 用最新声望剔除超限账号
excluded_by_prestige = set()
for account, (_, info) in account_latest_full_info.items():
    if info['声望'] > MAX_PRESTIGE:
        excluded_by_prestige.add(account)
if excluded_by_prestige:
    for day_dict in daily_accounts:
        for acc in excluded_by_prestige:
            day_dict.pop(acc, None)
    print(f"\n🔇 最新声望>{MAX_PRESTIGE} 剔除 {len(excluded_by_prestige)} 个账号")

# 用最新联盟剔除指定联盟账号
if EXCLUDE_ALLIANCES:
    excluded_by_alliance = set()
    for account, (_, info) in account_latest_full_info.items():
        if info['联盟'] in EXCLUDE_ALLIANCES:
            excluded_by_alliance.add(account)
    if excluded_by_alliance:
        for day_dict in daily_accounts:
            for acc in excluded_by_alliance:
                day_dict.pop(acc, None)
        print(f"🔇 排除联盟 {EXCLUDE_ALLIANCES} 剔除 {len(excluded_by_alliance)} 个账号")

# ==================== 汇总 ====================
total_days = len(daily_accounts)

if total_days == 0:
    print("❌ 没有有效数据可以分析")
    exit()

all_accounts = set()
for day_dict in daily_accounts:
    all_accounts.update(day_dict.keys())

miners = []

latest_n = min(N_LATEST_DAYS, total_days)

for account in all_accounts:
    mining_days = 0
    max_teams = 0
    latest_mining_info = None
    recent_mining_count = 0
    latest_mining_count = 0

    for i, day_dict in enumerate(daily_accounts):
        if account in day_dict:
            if day_dict[account]['is_mining']:
                mining_days += 1
                if i < MIN_RECENT_MINING:
                    recent_mining_count += 1
                if i < latest_n:
                    latest_mining_count += 1
            max_teams = max(max_teams, day_dict[account]['teams'])
            latest_mining_info = day_dict[account]['info']

    ratio = mining_days / total_days
    latest_ratio = latest_mining_count / latest_n if latest_n > 0 else 0

    if latest_ratio < LATEST_MIN:
        continue
    if ratio >= THRESHOLD or latest_ratio >= LATEST_THRESHOLD:
        is_inactive = recent_mining_count == 0

        # 使用最新完整信息（不受挖矿条件限制）
        if account in account_latest_full_info:
            latest_full_info = account_latest_full_info[account][1]
            final_nickname = latest_full_info['昵称']
            final_alliance = latest_full_info['联盟']
            final_furnace = latest_full_info['炉子']
            final_furnace_level = latest_full_info['炉子等级']
            final_prestige = latest_full_info['声望']
            final_coord = latest_full_info['坐标']
            final_shield = latest_full_info['罩子']
        else:
            # 兜底：使用最后的挖矿信息
            final_nickname = latest_mining_info['昵称'] if latest_mining_info else ''
            final_alliance = latest_mining_info['联盟'] if latest_mining_info else ''
            final_furnace = latest_mining_info['炉子'] if latest_mining_info else ''
            final_furnace_level = latest_mining_info['炉子等级'] if latest_mining_info else 0
            final_prestige = latest_mining_info['声望'] if latest_mining_info else 0
            final_coord = latest_mining_info['坐标'] if latest_mining_info else ''
            final_shield = latest_mining_info['罩子'] if latest_mining_info else ''

        miners.append({
            '账号': account,
            '昵称': final_nickname,
            '联盟': final_alliance,
            '炉子': final_furnace,
            '炉子等级': final_furnace_level,
            '声望': final_prestige,
            '坐标': final_coord,
            '罩子': final_shield,
            '最大采集队数': max_teams,
            '历史挖矿天数': f"{mining_days}/{total_days}",
            '最新挖矿天数': f"{latest_mining_count}/{latest_n}",
            '_ratio': ratio,
            '_mining_days': mining_days,
            '_latest_ratio': latest_ratio,
            '_is_inactive': is_inactive,
        })

df_miners = pd.DataFrame(miners)

if len(df_miners) > 0:
    df_miners = df_miners.sort_values(['_mining_days', '_latest_ratio', '声望'],
                                      ascending=[False, False, False])
    df_miners = df_miners.drop(columns=['_ratio', '_mining_days'])
    df_miners = df_miners.reset_index(drop=True)
    df_miners.index = df_miners.index + 1
    df_miners.index.name = '序号'

    green_count = (df_miners['_latest_ratio'] >= LATEST_THRESHOLD).sum()
    inactive_count = df_miners['_is_inactive'].sum()
    print(f"\n✅ 高频挖矿小号: {len(df_miners)} 个 (近期活跃: {green_count}, 退游矿工: {inactive_count})")
else:
    print(f"\n❌ 没有找到满足条件的玩家")

# ==================== 保存（带样式） ====================
# 清理旧输出文件
for old_file in glob.glob(os.path.join(CURRENT_DIR, '*_miner.xlsx')):
    try:
        os.remove(old_file)
        print(f"  🗑️ 已删除旧输出: {os.path.basename(old_file)}")
    except OSError:
        pass

output_file = f'{total_days}_miner.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    if len(df_miners) > 0:
        output_cols = ['账号', '昵称', '联盟', '声望', '坐标', '罩子',
                       '历史挖矿天数', '最新挖矿天数']
        # 罩子列中"无"显示为空
        df_miners['罩子'] = df_miners['罩子'].apply(lambda x: '' if str(x).strip() == '无' else x)
        df_miners[output_cols].to_excel(writer, sheet_name='高频挖矿小号', index=False)

        # 获取工作表并添加样式
        worksheet = writer.sheets['高频挖矿小号']

        # 基础样式
        base_font = Font(name='宋体', size=14)
        base_align = Alignment(horizontal='center', vertical='center')

        # 标记样式（主体不加粗）
        green_font = Font(name='宋体', size=14, color='008000')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        blue_font = Font(name='宋体', size=14, color='0066CC')
        blue_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

        # 遍历所有行，近期活跃标绿，退游矿工标蓝
        green_count = 0
        inactive_count = 0

        for row_idx in range(2, len(df_miners) + 2):
            miner_idx = row_idx - 2
            is_active = df_miners.iloc[miner_idx]['_latest_ratio'] >= LATEST_THRESHOLD
            is_inactive = df_miners.iloc[miner_idx]['_is_inactive']

            if is_active:
                green_count += 1
                for col_idx in range(1, len(output_cols) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = green_font
                    cell.fill = green_fill
                    cell.alignment = base_align
            elif is_inactive:
                inactive_count += 1
                for col_idx in range(1, len(output_cols) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = blue_font
                    cell.fill = blue_fill
                    cell.alignment = base_align
            else:
                for col_idx in range(1, len(output_cols) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = base_font
                    cell.alignment = base_align

        # 设置表头样式
        header_font_white = Font(name='宋体', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        for col_idx in range(1, len(output_cols) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 自动调整列宽（罩子列+20%）
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            header_val = str(column[0].value) if column[0].value else ''
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ''
                    # 中文字符按2个字符宽度计算
                    char_length = 0
                    for char in cell_value:
                        if '\u4e00' <= char <= '\u9fff' or '\u3000' <= char <= '\u303f':
                            char_length += 2
                        else:
                            char_length += 1
                    max_length = max(max_length, char_length)
                except:
                    pass
            width = max_length * 1.4 + 4
            if header_val == '罩子':
                width *= 1.2
            worksheet.column_dimensions[column_letter].width = width

        # 底部图例说明
        legend_row = len(df_miners) + 3
        note_font = Font(name='宋体', size=11, color='555555')
        legend_texts = [
            f'🟢 绿色 = 最新{N_LATEST_DAYS}天挖矿比例≥{int(LATEST_THRESHOLD * 100)}%（近期活跃）',
            f'🔵 蓝色 = 最近{MIN_RECENT_MINING}次采集无挖矿（退游矿工）',
            f'筛选条件：最新{N_LATEST_DAYS}天挖矿率≥{int(LATEST_MIN * 100)}%，且（历史挖矿率≥{int(THRESHOLD * 100)}% 或 最新{N_LATEST_DAYS}天挖矿率≥{int(LATEST_THRESHOLD * 100)}%）',
        ]
        for i, text in enumerate(legend_texts):
            cell = worksheet.cell(row=legend_row + i, column=1)
            cell.value = text
            cell.font = note_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            worksheet.merge_cells(start_row=legend_row + i, start_column=1,
                                  end_row=legend_row + i, end_column=len(output_cols))

        # 冻结首行
        worksheet.freeze_panes = 'A2'

print(f"\n📁 结果已保存: {output_file}")

if len(df_miners) > 0:
    print(f"\n🔥 高频挖矿小号（前20）:")
    for i, (_, row) in enumerate(df_miners.head(20).iterrows()):
        active_mark = f" 🟢活跃" if row['_latest_ratio'] >= LATEST_THRESHOLD else ""
        inactive_mark = " 🔵退游" if row['_is_inactive'] else ""
        print(
            f"  {i + 1}. [{row['联盟']}] {row['昵称']} | 声望{row['声望']} | {row['历史挖矿天数']} | {row['最新挖矿天数']}{active_mark}{inactive_mark}")

# ==================== 生成分析摘要 ====================
print(f"\n📊 分析摘要:")
print(f"  - 分析时间跨度: {len(sorted_dates)}天")
print(f"  - 处理文件数量: {len(selected_files)}个")
print(f"  - 发现矿工数量: {len(df_miners)}个")
if len(df_miners) > 0:
    print(f"  - 近期活跃(最新{N_LATEST_DAYS}天≥{int(LATEST_THRESHOLD*100)}%): {green_count}个")
    print(f"  - 退游矿工数量: {inactive_count}个（最近{MIN_RECENT_MINING}次无挖矿）")
    top_alliance = df_miners['联盟'].value_counts().head(1)
    if len(top_alliance) > 0:
        print(f"  - 最多矿工联盟: {top_alliance.index[0]} ({top_alliance.values[0]}个)")