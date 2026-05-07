import pandas as pd
import re
import os
import glob
import warnings
import matplotlib

matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.patheffects as pe
from datetime import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import matplotlib.font_manager as fm

warnings.filterwarnings('ignore')
for fp in ['C:/Windows/Fonts/msyh.ttc', 'C:/Windows/Fonts/simhei.ttf']:
    if os.path.exists(fp):
        fm.fontManager.addfont(fp)
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei']
plt.rcParams['axes.unicode_minus'] = False

# ==================== 🔧 配置区（按需修改） ====================

# 数据文件夹路径
DATA_FOLDER = 'data'

# 文件名匹配模式
FILE_PATTERN = '3957_*.xlsx'

# 【修改文件读取数量】总共读取的文件数量（包括最旧的1个 + 最新的N-1个）
READ_FILE_COUNT = 999  # 例如10表示：最旧1个 + 最新9个

# 图表筛选：声望大于多少才画图
MIN_PRESTIGE_FOR_PLOT = 000

# 最高声望限制（超过此值不显示，避免异常数据撑坏比例）
MAX_PRESTIGE = 20000

# 图表筛选条件
MIN_ALLIANCE_CHANGES = 0   # 最低联盟变化次数
MIN_NAME_CHANGES = 0      # 最低昵称变化次数
MIN_TOTAL_CHANGES = 0      # 最低联盟+昵称总变化次数

# 图表输出数量限制
TOP_N = 999

# 保护友方：True 时排除"出现过绿色联盟且从未出现过红色联盟"的玩家
PROTECT_GREEN_ONLY = True

# 排除低战无敌意：True 时排除声望低于 MIN_POWER 且从未出现过红色的玩家
FILTER_LOW_POWER = True
MIN_POWER = 2000

# ==================== 🎨 联盟颜色配置 ====================
RED_ALLIANCES = ['FFF', '666', 'OoO', 'SSR']
RED_COLOR = '#E74C3C'

GREEN_ALLIANCES = ['999', 'ann', 'xxx']
GREEN_COLOR = '#27AE60'

BLUE_ALLIANCES = ['KFC']
BLUE_COLOR = '#3498DB'

DEFAULT_COLOR = '#000000'

ALLIANCE_COLORS = {}
for a in RED_ALLIANCES:
    ALLIANCE_COLORS[a.upper()] = RED_COLOR
for a in GREEN_ALLIANCES:
    ALLIANCE_COLORS[a.upper()] = GREEN_COLOR
for a in BLUE_ALLIANCES:
    ALLIANCE_COLORS[a.upper()] = BLUE_COLOR

LEGEND_LABELS = {
    RED_COLOR: '/'.join(RED_ALLIANCES),
    GREEN_COLOR: '/'.join(GREEN_ALLIANCES),
    BLUE_COLOR: '/'.join(BLUE_ALLIANCES),
}


def get_alliance_color(alliance):
    if pd.isna(alliance) or alliance == '' or alliance == '-':
        return DEFAULT_COLOR
    return ALLIANCE_COLORS.get(alliance.strip().upper(), DEFAULT_COLOR)


# ==================== 加载豁免账号 ====================
def load_protected_accounts(folder_path):
    """从data文件夹加载change_lm&name_protect.xlsx豁免账号"""
    protect_file = os.path.join(folder_path, 'change_lm&name_protect.xlsx')

    if not os.path.exists(protect_file):
        print(f"ℹ️  未找到豁免文件: {protect_file}，将不豁免任何账号")
        return []

    try:
        protect_df = pd.read_excel(protect_file, dtype=str)

        # 检查必要的列
        if '账号' not in protect_df.columns:
            print(f"⚠️  豁免文件格式错误，需要包含'账号'列")
            return []

        # 提取账号列表
        protected = []
        for _, row in protect_df.iterrows():
            account = row['账号']
            remark = row.get('备注', '')

            try:
                account = int(float(account))
                protected.append(account)
                if pd.notna(remark) and str(remark).strip() != '':
                    print(f"  🛡️  豁免: {account} ({remark})")
                else:
                    print(f"  🛡️  豁免: {account}")
            except (ValueError, TypeError):
                print(f"  ⚠️  无法解析账号: {account}")
                continue

        return protected

    except Exception as e:
        print(f"❌ 读取豁免文件失败: {e}")
        return []


# ==================== 查找文件 ====================
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER_PATH = os.path.join(CURRENT_DIR, DATA_FOLDER)

print("🛡️ 加载豁免账号...")
EXEMPT_ACCOUNTS = load_protected_accounts(FOLDER_PATH)
print(f"  共豁免 {len(EXEMPT_ACCOUNTS)} 个账号\n")

all_files = glob.glob(os.path.join(FOLDER_PATH, FILE_PATTERN))

# 过滤掉备注和豁免文件
all_files = [f for f in all_files if 'miner_remark' not in os.path.basename(f).lower()
             and 'change_lm' not in os.path.basename(f).lower()
             and 'name_protect' not in os.path.basename(f).lower()]

print(f"找到 {len(all_files)} 个数据文件")


# ==================== 提取文件时间信息 ====================
def extract_datetime_from_filename(filename):
    """
    从文件名中提取日期时间信息
    支持格式：
    - 3957_0428.xlsx (一天一次)
    - 3957_0428a.xlsx, 3957_0428b.xlsx (一天多次)
    返回 (日期对象, 显示字符串, 排序元组)
    """
    match = re.search(r'3957_(\d{4})([a-z]?)\.xlsx', filename)
    if match:
        date_str = match.group(1)  # 0428
        suffix = match.group(2)  # a/b/c 或空字符串

        # 解析日期
        try:
            month = int(date_str[:2])
            day = int(date_str[2:])
            # 假设2026年（可根据需要修改）
            date_obj = datetime(2026, month, day)

            # 显示字符串
            if suffix:
                display_str = f"{date_str}{suffix}"  # 0428a
            else:
                display_str = f"{date_str}"  # 0428

            # 排序键：同一天内 a=1, b=2, c=3..., 无后缀=0（最早）
            suffix_order = ord(suffix) - ord('a') + 1 if suffix else 0

            return date_obj, display_str, suffix_order
        except ValueError:
            pass

    return None, None, 0


# 提取所有文件的时间信息
file_infos = []
for f in all_files:
    basename = os.path.basename(f)
    date_obj, display_str, suffix_order = extract_datetime_from_filename(basename)
    if date_obj and display_str:
        # 排序键：(日期, 后缀顺序)
        sort_key = (date_obj, suffix_order)
        file_infos.append((f, date_obj, display_str, sort_key))
    else:
        print(f"  ⚠️ 无法提取日期: {basename}")

if len(file_infos) == 0:
    print("❌ 无有效文件")
    exit()

# 按时间排序
file_infos.sort(key=lambda x: x[3])

print(f"共 {len(file_infos)} 个有效文件，时间范围：{file_infos[0][2]} ~ {file_infos[-1][2]}")

# ==================== 选择文件：最旧1个 + 最新N-1个 ====================
if len(file_infos) <= READ_FILE_COUNT:
    selected_files = file_infos
    print(f"✅ 文件总数({len(file_infos)})≤读取限制({READ_FILE_COUNT})，全部读取")
else:
    # 最旧的文件（第一个）
    oldest = file_infos[0]

    # 最新的READ_FILE_COUNT-1个文件
    newest = file_infos[-(READ_FILE_COUNT - 1):]

    # 合并并去重（按文件路径）
    selected_set = {oldest[0]: oldest}
    for f in newest:
        selected_set[f[0]] = f

    selected_files = sorted(selected_set.values(), key=lambda x: x[3])

    print(f"✅ 选择策略：最旧1个 + 最新{READ_FILE_COUNT - 1}个 = 共{len(selected_files)}个文件")

print("\n📂 选中的文件:")
for f, date_obj, display_str, _ in selected_files:
    print(f"  📄 {os.path.basename(f)} → {display_str}")

# 构建时间映射
time_order = [info[2] for info in selected_files]  # 显示字符串列表
time_dates = [info[1] for info in selected_files]  # 日期对象列表

# ==================== 计算累积天数 ====================
date_diffs = []
for i in range(len(time_dates)):
    if i == 0:
        date_diffs.append(0)
    else:
        diff = (time_dates[i] - time_dates[i - 1]).days
        # 同一天的文件（后缀不同）间隔设为0.8
        if diff == 0:
            date_diffs.append(0.8)
        elif diff == 1:
            date_diffs.append(1.0)
        elif diff == 2:
            date_diffs.append(1.2)
        elif diff == 3:
            date_diffs.append(1.4)
        else:
            date_diffs.append(1.5)

cumulative_days = np.cumsum(date_diffs)
total_days_span = cumulative_days[-1] - cumulative_days[0]

print(f"\n📊 时间跨度: {total_days_span:.1f} (累计)")


# ==================== 读取数据 ====================
def safe_to_numeric(series, default_value=0):
    result = pd.to_numeric(series, errors='coerce')
    return result.fillna(default_value).astype(float)


account_timeline = {}

for file_path, date_obj, display_str, _ in selected_files:
    try:
        df = pd.read_excel(file_path, sheet_name=0, dtype=str)
    except Exception as e:
        print(f"  ❌ 读取失败 {os.path.basename(file_path)}: {e}")
        continue

    df['账号'] = safe_to_numeric(df['账号'])
    df['声望'] = safe_to_numeric(df['声望'])

    for _, row in df.iterrows():
        account = int(row['账号'])
        if account == 0:
            continue
        if account in EXEMPT_ACCOUNTS:
            continue

        alliance = str(row['联盟']).strip() if pd.notna(row['联盟']) else '-'
        name = str(row['昵称']).strip() if pd.notna(row['昵称']) else ''
        prestige = min(int(row['声望']), MAX_PRESTIGE)

        if account not in account_timeline:
            account_timeline[account] = {}

        account_timeline[account][display_str] = {
            '昵称': name,
            '联盟': alliance,
            '声望': prestige,
        }

print(f"\n追踪 {len(account_timeline)} 个账号（已排除 {len(EXEMPT_ACCOUNTS)} 个豁免）")

# 排除最新文件中未出现的玩家（可能已退游）
latest_time = time_order[-1]
active_accounts = {a for a, tl in account_timeline.items() if latest_time in tl}
removed = len(account_timeline) - len(active_accounts)
if removed > 0:
    print(f"  排除 {removed} 个最新未出现账号（可能已退游）")
    account_timeline = {k: v for k, v in account_timeline.items() if k in active_accounts}

# ==================== 分析变化 ====================
change_records = []

for account, timeline in account_timeline.items():
    sorted_times = sorted(timeline.keys(), key=lambda x: time_order.index(x) if x in time_order else 999)
    if len(sorted_times) < 2:
        continue

    name_changes = []
    alliance_changes = []

    prev_name = None
    prev_alliance = None

    for t in sorted_times:
        current_name = timeline[t]['昵称']
        current_alliance = timeline[t]['联盟']

        if prev_name is not None and current_name != prev_name:
            name_changes.append((t, prev_name, current_name))

        if prev_alliance is not None and current_alliance != prev_alliance:
            if prev_alliance != '-' and current_alliance != '-':
                alliance_changes.append((t, prev_alliance, current_alliance))
            elif prev_alliance != '-' and current_alliance == '-':
                pass
            elif prev_alliance == '-' and current_alliance != '-':
                last_real_alliance = None
                for pt in reversed(sorted_times[:sorted_times.index(t)]):
                    if timeline[pt]['联盟'] != '-':
                        last_real_alliance = timeline[pt]['联盟']
                        break
                if last_real_alliance and last_real_alliance != current_alliance:
                    alliance_changes.append((t, last_real_alliance, current_alliance))

        prev_name = current_name
        prev_alliance = current_alliance

    total_changes = len(name_changes) + len(alliance_changes)

    if total_changes > 0:
        earliest = timeline[sorted_times[0]]
        latest = timeline[sorted_times[-1]]

        name_change_summary = '→'.join([name_changes[0][1]] + [new for _, _, new in name_changes]) if name_changes else ''
        if alliance_changes:
            alliance_chain = [alliance_changes[0][1]] + [new for _, _, new in alliance_changes]
            alliance_change_summary = '→'.join(alliance_chain)
        else:
            alliance_change_summary = ''

        change_dates = set()
        for t_data in name_changes:
            change_dates.add(t_data[0])
        for t_data in alliance_changes:
            change_dates.add(t_data[0])
        if len(sorted_times) > 0:
            change_dates.add(sorted_times[0])
            change_dates.add(sorted_times[-1])

        change_records.append({
            '账号': account,
            '昵称': earliest['昵称'],
            '联盟': earliest['联盟'],
            '声望': latest['声望'],
            '昵称变化次数': len(name_changes),
            '联盟变化次数': len(alliance_changes),
            '总变化次数': total_changes,
            '昵称变化': name_change_summary,
            '联盟变化': alliance_change_summary,
            '有联盟变化': len(alliance_changes) > 0,
            '_timeline': timeline,
            '_sorted_times': sorted_times,
            '_name_changes': name_changes,
            '_alliance_changes': alliance_changes,
            '_change_dates': change_dates,
        })

df_changes = pd.DataFrame(change_records)

if len(df_changes) > 0:
    df_alliance_changed = df_changes[df_changes['有联盟变化']].sort_values(
        ['联盟变化次数', '声望'], ascending=[False, False])
    df_name_only = df_changes[~df_changes['有联盟变化']].sort_values(
        ['昵称变化次数', '声望'], ascending=[False, False])
    df_changes = pd.concat([df_alliance_changed, df_name_only], ignore_index=True)
else:
    print("❌ 无变化")
    exit()

# ==================== 保存Excel ====================
output_excel = '昵称联盟变化分析.xlsx'
output_cols = ['账号', '昵称', '联盟', '声望', '昵称变化', '联盟变化']
left_align_cols = {'昵称变化', '联盟变化'}  # 文本列左对齐

body_font = Font(name='宋体', size=14)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    # 工作表1: 变化总表（按最新声望降序）
    df_all = df_changes.sort_values('声望', ascending=False).reset_index(drop=True)
    df_all[output_cols].to_excel(writer, sheet_name='变化总表', index=False)

    # 工作表2: 联盟变化排名（仅保留有联盟变化者）
    df_sheet1 = df_changes[df_changes['有联盟变化']].sort_values(
        ['联盟变化次数', '声望'], ascending=[False, False])
    df_sheet1[output_cols].to_excel(writer, sheet_name='联盟变化排名', index=False)

    # 工作表3: 昵称变化排名（仅保留有昵称变化者）
    df_sheet2 = df_changes[df_changes['昵称变化次数'] > 0].sort_values(
        ['昵称变化次数', '声望'], ascending=[False, False])
    df_sheet2[output_cols].to_excel(writer, sheet_name='昵称变化排名', index=False)

    # ========== 样式 ==========
    for sheet_name in ['变化总表', '联盟变化排名', '昵称变化排名']:
        ws = writer.sheets[sheet_name]

        # 表头
        for col_idx in range(1, len(output_cols) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(name='宋体', size=14, bold=True)
            cell.alignment = center_align

        # 数据行
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(output_cols) + 1):
                col_name = output_cols[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.alignment = left_align if col_name in left_align_cols else center_align

        # 自动列宽
        for col_idx in range(1, len(output_cols) + 1):
            max_char = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                char_len = 0
                for ch in val:
                    if '一' <= ch <= '鿿' or '　' <= ch <= '〿':
                        char_len += 2
                    else:
                        char_len += 1
                max_char = max(max_char, char_len)
            ws.column_dimensions[col_letter].width = max_char * 1.4 + 4

        # 冻结首行
        ws.freeze_panes = 'A2'

print(f"\n📁 {output_excel}")
print(f"   📋 变化总表 - {len(df_all)}人（按最新声望降序）")
print(f"   📋 联盟变化排名 - {len(df_sheet1)}人（仅联盟变化）")
print(f"   📋 昵称变化排名 - {len(df_sheet2)}人（仅昵称变化）")


# ==================== 通用绘图函数（优化版） ====================
def draw_timeline(df_plot_full, output_name):
    """绘制优化后的时序图——合并联盟+昵称变化、声望降序"""
    df_plot = df_plot_full.head(TOP_N).copy()
    if len(df_plot) == 0:
        print(f"⚠️ 无符合条件的玩家，跳过 {output_name}")
        return

    n_players = len(df_plot)

    # 计算最长标签文字，用于边距调整
    max_left_text = ""
    max_right_text = ""
    for _, row in df_plot.iterrows():
        tl = row['_timeline']
        st = row['_sorted_times']
        ft = st[0]
        lt = st[-1]
        fa = tl[ft]['联盟']
        fn = tl[ft]['昵称']
        lt_txt = f"({fa}){fn}" if fa != '-' else fn
        if len(lt_txt) > len(max_left_text):
            max_left_text = lt_txt
        la = tl[lt]['联盟']
        ln = tl[lt]['昵称']
        rt_txt = f"({la}){ln}" if la != '-' else ln
        if len(rt_txt) > len(max_right_text):
            max_right_text = rt_txt

    # --- 动态尺寸（略微加大行高和宽度）---
    fig_height = max(9, n_players * 0.5)
    fig_width = max(13, min(44, total_days_span * 0.55))

    # --- 动态字体（提高最小字号）---
    if n_players > 25:
        label_fontsize = 6.5
        point_size = 20
        line_width = 1.8
        stat_fontsize = 6
    elif n_players > 15:
        label_fontsize = 7.5
        point_size = 24
        line_width = 2.2
        stat_fontsize = 7
    else:
        label_fontsize = 9
        point_size = 30
        line_width = 2.8
        stat_fontsize = 8

    bar_max_length = total_days_span * 0.06

    # 声望条起始偏移：按右侧最长标签计算，避免重叠
    right_label_width_pt = len(max_right_text) * label_fontsize * 0.85
    bar_offset_mult = max(0.14, (12 + right_label_width_pt) / 72 * 1.4 / fig_width)

    fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=150)

    all_prestige = [row['声望'] for _, row in df_plot.iterrows()]
    max_prestige = max(all_prestige) if all_prestige else 1
    min_prestige = min(all_prestige) if all_prestige else 0
    prestige_range = max_prestige - min_prestige if max_prestige > min_prestige else 1

    # --- ① 交替行背景（提升行间区分度）---
    for i in range(n_players):
        if i % 2 == 1:
            ax.axhspan(i + 0.5, i + 1.5, facecolor='#F4F6F8', zorder=0, alpha=0.6)

    # --- ② 水平辅助网格线 ---
    ax.grid(axis='y', alpha=0.15, linestyle='-', linewidth=0.5, zorder=0)
    ax.grid(axis='x', alpha=0.2, linestyle='--', linewidth=0.5, zorder=0)

    # 最小时间间距（用于水平重叠检测）
    min_gap = min(cumulative_days[k] - cumulative_days[k-1] for k in range(1, len(cumulative_days))) if len(cumulative_days) > 1 else 1.0

    for i, (_, row) in enumerate(df_plot.iterrows()):
        y_pos = i + 1
        timeline = row['_timeline']
        sorted_times = row['_sorted_times']
        change_dates = row['_change_dates']

        # --- ③ 时间线段（加粗、提升对比度）---
        for j in range(len(sorted_times) - 1):
            t_curr = sorted_times[j]
            t_next = sorted_times[j + 1]

            if t_curr in time_order and t_next in time_order:
                x_curr = cumulative_days[time_order.index(t_curr)]
                x_next = cumulative_days[time_order.index(t_next)]
            else:
                continue

            alliance_curr = timeline[t_curr]['联盟']
            if alliance_curr == '-':
                color = '#AAAAAA'
                alpha = 0.5
                linestyle = ':'
            else:
                color = get_alliance_color(alliance_curr)
                alpha = 0.75
                linestyle = '-'

            ax.plot([x_curr, x_next], [y_pos, y_pos], color=color, linewidth=line_width,
                    alpha=alpha, linestyle=linestyle, zorder=2)

        # --- ④ 变化点标记 ---
        for t in sorted_times:
            if t not in change_dates or t not in time_order:
                continue

            x = cumulative_days[time_order.index(t)]
            alliance = timeline[t]['联盟']

            if alliance == '-':
                ax.scatter(x, y_pos, s=point_size, facecolors='none', edgecolors='#888888',
                           linewidths=1.5, alpha=0.7, zorder=4)
            else:
                ax.scatter(x, y_pos, s=point_size, c=get_alliance_color(alliance), alpha=0.9,
                           edgecolors='white', linewidths=0.8, zorder=4)

        # --- ⑤ 标注变化（联盟统一在上/昵称统一在下，只标变化后信息）---
        prev_name = timeline[sorted_times[0]]['昵称']
        prev_alliance = timeline[sorted_times[0]]['联盟']

        for j, t in enumerate(sorted_times):
            if j == 0 or j == len(sorted_times) - 1 or t not in time_order:
                continue

            x = cumulative_days[time_order.index(t)]
            current_name = timeline[t]['昵称']
            current_alliance = timeline[t]['联盟']

            has_alliance = (current_alliance != prev_alliance and
                           prev_alliance != '-' and current_alliance != '-')
            has_name = (current_name != prev_name)

            if not has_alliance and not has_name:
                prev_name = current_name
                prev_alliance = current_alliance
                continue

            # 联盟变化标注在线之上（只标新联盟）
            if has_alliance:
                ax.annotate(current_alliance,
                            xy=(x, y_pos), xytext=(0, 6),
                            textcoords='offset points', fontsize=label_fontsize, fontweight='bold',
                            ha='center', va='bottom',
                            color=get_alliance_color(current_alliance),
                            path_effects=[pe.withStroke(linewidth=2.5, foreground='white')],
                            zorder=10)

            # 昵称变化标注在线之下（只标新昵称）
            if has_name:
                ax.annotate(current_name,
                            xy=(x, y_pos), xytext=(0, -6),
                            textcoords='offset points', fontsize=label_fontsize, fontweight='bold',
                            ha='center', va='top', color='#666666',
                            path_effects=[pe.withStroke(linewidth=2.5, foreground='white')],
                            zorder=10)

            prev_name = current_name
            prev_alliance = current_alliance

        # --- ⑥ 左侧：起始信息（带边框圆角背景）---
        first_t = sorted_times[0]
        if first_t in time_order:
            x_first = cumulative_days[time_order.index(first_t)]
            first_alliance = timeline[first_t]['联盟']
            first_name = timeline[first_t]['昵称']
            first_color = get_alliance_color(first_alliance) if first_alliance != '-' else '#888888'

            label_text = f"({first_alliance}){first_name}" if first_alliance != '-' else first_name

            ax.annotate(label_text,
                        xy=(x_first, y_pos),
                        xytext=(-8, 0),
                        textcoords='offset points',
                        fontsize=label_fontsize, color=first_color,
                        ha='right', va='center', fontweight='bold',
                        zorder=8)

        # --- ⑦ 右侧：最新状态（在最后一点右侧）+ 声望条 ---
        last_t = sorted_times[-1]
        if last_t in time_order:
            x_last = cumulative_days[time_order.index(last_t)]
            prestige = row['声望']

            # 声望条长度（严格按比值）
            if max_prestige > 0:
                bar_length = bar_max_length * prestige / max_prestige
            else:
                bar_length = bar_max_length * 0.3

            # 声望条位置（在最新状态右侧，避开右侧标签）
            bar_start_x = x_last + total_days_span * bar_offset_mult

            # 声望条
            bar_color = get_alliance_color(timeline[sorted_times[-1]]['联盟'])
            if timeline[sorted_times[-1]]['联盟'] == '-':
                bar_color = '#999999'

            ax.barh(y_pos, bar_length, left=bar_start_x, height=0.35,
                    color=bar_color, alpha=0.8, edgecolor='white', linewidth=0.5, zorder=3)
            ax.text(bar_start_x + bar_length + total_days_span * 0.008, y_pos,
                    f"{prestige:,}",
                    fontsize=stat_fontsize, va='center', ha='left', color='#000', fontweight='bold')

            # 最新状态（在最后一个点右侧，与初始状态对称）
            last_name = timeline[last_t]['昵称']
            last_alliance = timeline[last_t]['联盟']
            if last_name:
                last_label = f"({last_alliance}){last_name}" if last_alliance != '-' else last_name
                last_color = get_alliance_color(last_alliance) if last_alliance != '-' else '#888888'
                ax.annotate(last_label,
                            xy=(x_last, y_pos),
                            xytext=(8, 0), textcoords='offset points',
                            fontsize=label_fontsize, color=last_color,
                            ha='left', va='center', fontweight='bold',
                            zorder=8)

    # --- ⑧ 表头行（初始 | 声望 | 最新） ---
    hdr_fs = max(11, stat_fontsize + 5)
    x_last_global = cumulative_days[-1]
    hdr_bar_x = x_last_global + total_days_span * bar_offset_mult
    hdr_y = 0.62

    ax.annotate('初始状态', xy=(cumulative_days[0], hdr_y), xytext=(-8, 0),
                textcoords='offset points', fontsize=hdr_fs, va='center', ha='right',
                fontweight='bold', color='#000', alpha=0.85, zorder=1, fontfamily='STXingkai')
    ax.text(hdr_bar_x, hdr_y, '声望', fontsize=hdr_fs, va='center', ha='left',
            fontweight='bold', color='#000', alpha=0.85, zorder=1, fontfamily='STXingkai')
    ax.annotate('最新状态', xy=(x_last_global, hdr_y), xytext=(8, 0),
                textcoords='offset points', fontsize=hdr_fs, va='center', ha='left',
                fontweight='bold', color='#000', alpha=0.85, zorder=1, fontfamily='STXingkai')

    # --- ⑨ 标题 ---
    date_range = f"{time_order[0]} ~ {time_order[-1]}"
    chart_title = f"玩家变化时序图 | {date_range} | 声望>{MIN_PRESTIGE_FOR_PLOT}"
    ax.set_title(chart_title, fontsize=14, fontweight='bold', pad=18, color='#222',
                 fontfamily='STXingkai')

    # --- ⑩ X轴优化（标签自动跳过避免拥挤）---
    ax.set_xticks(cumulative_days)

    max_labels = max(10, int(fig_width / 1.8))
    if len(time_order) > max_labels:
        step = len(time_order) // max_labels + 1
        for idx, label in enumerate(ax.get_xticklabels()):
            if idx % step != 0:
                label.set_visible(False)

    if total_days_span > 30:
        xtick_fontsize = 7
        rotation = 40
    elif total_days_span > 15:
        xtick_fontsize = 8.5
        rotation = 25
    else:
        xtick_fontsize = 10
        rotation = 0
    ax.set_xticklabels(time_order, fontsize=xtick_fontsize, rotation=rotation)
    ax.annotate('探查时间', xy=(0.5, 0), xycoords='axes fraction',
                xytext=(0, -28), textcoords='offset points',
                fontsize=13, fontweight='bold', ha='center', va='top',
                fontfamily='STXingkai')

    # --- ⑪ 样式美化 ---
    ax.spines['top'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_color('#000000')
    ax.tick_params(left=False, labelleft=False)
    ax.set_ylabel('')

    # 增加边距让两侧内容有呼吸空间
    char_margin = label_fontsize * 0.0012
    left_margin = total_days_span * max(0.10, len(max_left_text) * char_margin)
    right_margin = total_days_span * max(0.28, len(max_right_text) * char_margin)
    ax.set_xlim(cumulative_days[0] - left_margin, cumulative_days[-1] + right_margin)
    ax.set_ylim(n_players + 1.0, 0.5)

    # --- ⑫ 图例（水平一行，放到图表下方）---
    from matplotlib.lines import Line2D
    legend_elements = []
    if RED_ALLIANCES:
        legend_elements.append(Line2D([0], [0], marker='o', color='w', markerfacecolor=RED_COLOR,
                                      markersize=8, label=LEGEND_LABELS[RED_COLOR]))
    if GREEN_ALLIANCES:
        legend_elements.append(Line2D([0], [0], marker='o', color='w', markerfacecolor=GREEN_COLOR,
                                      markersize=8, label=LEGEND_LABELS[GREEN_COLOR]))
    if BLUE_ALLIANCES:
        legend_elements.append(Line2D([0], [0], marker='o', color='w', markerfacecolor=BLUE_COLOR,
                                      markersize=8, label=LEGEND_LABELS[BLUE_COLOR]))
    legend_elements.append(Line2D([0], [0], marker='o', color='w', markerfacecolor=DEFAULT_COLOR,
                                  markersize=8, label='其他联盟'))
    legend_elements.append(Line2D([0], [0], marker='o', color='w',
                                  markerfacecolor='none', markeredgecolor='#888888',
                                  markersize=8, markeredgewidth=1.5, label='无联盟'))

    ax.legend(handles=legend_elements, fontsize=7,
              loc='upper center', bbox_to_anchor=(0.5, 0),
              borderaxespad=88/7,
              frameon=False, ncol=5, columnspacing=1.2)

    # --- ⑬ 底部输出限制（固定点数间距，不受图高影响） ---
    ax.annotate(f'战力区间[{MIN_PRESTIGE_FOR_PLOT}, {MAX_PRESTIGE}]',
                xy=(0.5, 0), xycoords='axes fraction',
                xytext=(0, -52), textcoords='offset points',
                fontsize=7, ha='center', va='top', color='#000',
                fontfamily='Microsoft YaHei')
    ax.annotate('规则1:曾加入绿色联盟且无红色历史的玩家不显示',
                xy=(0.5, 0), xycoords='axes fraction',
                xytext=(0, -64), textcoords='offset points',
                fontsize=7, ha='center', va='top', color='#000',
                fontfamily='Microsoft YaHei')
    ax.annotate(f'规则2:战力低于{MIN_POWER}且无红色历史的玩家不显示',
                xy=(0.5, 0), xycoords='axes fraction',
                xytext=(0, -76), textcoords='offset points',
                fontsize=7, ha='center', va='top', color='#000',
                fontfamily='Microsoft YaHei')

    plt.tight_layout()
    fig.savefig(output_name, dpi=150, bbox_inches='tight', facecolor='white')
    print(f"✅ {output_name} 已保存 ({n_players}/{len(df_plot_full)}个玩家)")


# ==================== 准备绘图数据（合并为一张图，声望降序） ====================
df_plot = df_changes[df_changes['声望'] > MIN_PRESTIGE_FOR_PLOT].copy()
df_plot = df_plot[(df_plot['联盟变化次数'] >= MIN_ALLIANCE_CHANGES) &
                  (df_plot['昵称变化次数'] >= MIN_NAME_CHANGES) &
                  (df_plot['总变化次数'] >= MIN_TOTAL_CHANGES)]

# 保护友方：排除仅出现过绿色联盟的玩家
if PROTECT_GREEN_ONLY:
    def _has_green_only(row):
        timeline = row['_timeline']
        green_set = {a.upper() for a in GREEN_ALLIANCES}
        red_set = {a.upper() for a in RED_ALLIANCES}
        has_green = False
        for data in timeline.values():
            alliance = data['联盟'].strip().upper() if pd.notna(data['联盟']) and data['联盟'] != '-' else ''
            if alliance in red_set:
                return False
            if alliance in green_set:
                has_green = True
        return has_green

    exclude_mask = df_plot.apply(_has_green_only, axis=1)
    excluded = exclude_mask.sum()
    if excluded > 0:
        print(f"  🟢 保护友方：排除 {excluded} 个仅出现在绿色联盟的玩家")
        df_plot = df_plot[~exclude_mask]

# 排除低战无敌意玩家
if FILTER_LOW_POWER:
    def _is_low_and_safe(row):
        if row['声望'] >= MIN_POWER:
            return False
        red_set = {a.upper() for a in RED_ALLIANCES}
        for data in row['_timeline'].values():
            alliance = data['联盟'].strip().upper() if pd.notna(data['联盟']) and data['联盟'] != '-' else ''
            if alliance in red_set:
                return False
        return True

    low_mask = df_plot.apply(_is_low_and_safe, axis=1)
    excluded_low = low_mask.sum()
    if excluded_low > 0:
        print(f"  🔻 排除 {excluded_low} 个低战无敌意玩家（声望<{MIN_POWER}）")
        df_plot = df_plot[~low_mask]

df_plot = df_plot.sort_values('声望', ascending=False).reset_index(drop=True)

# ==================== 生成一张图 ====================
if len(df_plot) > 0:
    draw_timeline(df_plot, '玩家变化时序图.png')
else:
    print(f"⚠️ 无声望>{MIN_PRESTIGE_FOR_PLOT}的玩家")

print(f"\n✅ 分析完成！")